from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy
import pandas as pd
import os


def monthly_to_excel(excel_file):
    # 엑셀 파일 열기
    wb = load_workbook(excel_file)
    ws = wb.active

    print(f"파일 열기: {excel_file}")
    print("작업 시작")

    # Step 1: 1, 2번 행 삭제
    print("1단계: 1, 2번 행 삭제 중...")
    ws.delete_rows(1, 2)

    # Step 2: A~E 열 삭제
    print("2단계: A~E 열 삭제 중...")
    ws.delete_cols(1, 5)

    # Step 3: B~Y 열 삭제
    print("3단계: B~Y 열 삭제 중...")
    ws.delete_cols(2, 24)

    # Step 4: 필터 추가
    print("4단계: 필터 추가 중...")
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Step 5: A열에서 "FashionGo_공통" 필터링 및 나머지 행 삭제
    print("5단계: A열 필터링 - 'FashionGo_공통' 만 유지 중...")
    rows_to_delete = []
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=1).value
        if cell_value != "FashionGo_공통":
            rows_to_delete.append(row_idx)

    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx, 1)

    # Step 6: 데이터를 pandas DataFrame으로 읽기
    print("6단계: 피벗 테이블 생성 중...")
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)
    
    # 첫 번째 행(헤더) 가져오기
    headers = [cell.value for cell in ws[1]]
    df = pd.DataFrame(data, columns=headers)
    
    # 데이터 타입 확인 및 변환
    # 숫자 필드는 float로 변환
    numeric_columns = ['개인 서비스별 실적(시간)', '인월(서비스별 실적/합계실적)(시간)']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Step 7: 피벗 테이블 생성
    # 행: 프로젝트명, 사원이름
    # 값: 개인 서비스별 실적(시간), 인월(서비스별 실적/합계실적)(시간)
    pivot_table = df.pivot_table(
        index=['프로젝트명', '사원이름'],
        values=['개인 서비스별 실적(시간)', '인월(서비스별 실적/합계실적)(시간)'],
        aggfunc='sum',
        margins=True,
        margins_name='합계'
    )
    
    # 디버깅: 피벗 테이블 인덱스 확인
    print("\n=== 피벗 테이블 인덱스 ===")
    for idx, index_val in enumerate(pivot_table.index):
        print(f"{idx}: {index_val}")
    print("=========================\n")
    
    # Step 8: 피벗 테이블을 새 시트에 작성
    pivot_ws = wb.create_sheet("피벗_분석")
    
    # 헤더 작성 (행 인덱스 헤더)
    pivot_ws.cell(row=1, column=1, value='행 레이블')
    
    # 값 헤더 작성
    for col_idx, col_name in enumerate(pivot_table.columns, 2):
        pivot_ws.cell(row=1, column=col_idx, value=col_name)
    
    # 현재 행 번호
    current_row = 2
    total_row = None  # 총합계 행 번호 저장
    
    # 피벗 테이블의 모든 인덱스를 리스트로 변환
    pivot_indices = list(pivot_table.index)
    
    # 피벗 테이블 데이터 쓰기
    for index_value, row_data in pivot_table.iterrows():
        # 행 인덱스 작성
        if isinstance(index_value, tuple):
            project = index_value[0]
            employee = index_value[1]
            
            # 마지막 행(전체 합계)인지 확인 - '합계' in 인덱스 확인
            if employee == '합계' and project == '합계':
                pivot_ws.cell(row=current_row, column=1, value='총합계')
                total_row = current_row  # 총합계 행 번호 저장
                print(f"✅ 총합계 행 저장됨: {current_row}행 ({project}, {employee})")
            elif employee == '합계':
                # 프로젝트별 합계
                pivot_ws.cell(row=current_row, column=1, value=project)
                print(f"프로젝트별 합계: {current_row}행 ({project})")
            else:
                # 개별 사원
                pivot_ws.cell(row=current_row, column=1, value=employee)
        else:
            pivot_ws.cell(row=current_row, column=1, value=index_value)
        
        # 데이터 값 작성
        for col_idx, value in enumerate(row_data, 2):
            pivot_ws.cell(row=current_row, column=col_idx, value=value)
        
        current_row += 1
    
    # Step 9: 총합계 행 볼드 처리
    if total_row:
        print(f"\n✅ 총합계 행 볼드 처리 중: {total_row}행")
        for col_idx in range(1, pivot_ws.max_column + 1):
            cell = pivot_ws.cell(row=total_row, column=col_idx)
            cell.font = Font(bold=True)
    else:
        print("\n⚠️  총합계 행을 찾을 수 없습니다.")
    
    # 파일 저장 (기존 파일이 있으면 삭제)
    output_file = "output/monthly_output.xlsx"
    os.makedirs("output", exist_ok=True)
    
    # 기존 파일이 열려있으면 닫기
    if os.path.exists(output_file):
        try:
            os.remove(output_file)
        except PermissionError:
            print("⚠️  기존 파일이 열려있습니다. 먼저 닫은 후 다시 실행하세요.")
            return
    
    wb.save(output_file)
    print(f"\n✅ 완료! 파일: {output_file}")