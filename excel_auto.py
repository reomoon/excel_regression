from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

# 작업할 엑셀 파일 이름
excel_file = "upload/reg_upload.xlsx"

# 엑셀 파일 열기
wb = load_workbook(excel_file)
ws = wb.active

print("작업 시작...")

# 1단계: 첫 줄 A~O 복사 후 3번째 줄에 붙여넣기
print("1단계: 첫 줄 복사 중...")
for col in range(1, 16):  # A부터 O까지
    source_cell = ws.cell(row=1, column=col)
    target_cell = ws.cell(row=3, column=col)
    target_cell.value = source_cell.value
    # 셀 스타일 복사 (색, 테두리 등)
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

# 2단계: 1, 2번 행 삭제
print("2단계: 1, 2번 행 삭제 중...")
ws.delete_rows(1, 2)

# 3단계: B 열 삭제
print("3단계: B 열 삭제 중...")
ws.delete_cols(2, 1)  # B는 2번째 열

# 4단계: AC 열 복사 후 D 열 앞에 삽입
print("4단계: AC 열 복사 중...")
ac_col_index = 29
ac_col_after_delete = ac_col_index - 1

# AC 열 데이터 저장
ac_data = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=ac_col_after_delete, max_col=ac_col_after_delete):
    for cell in row:
        ac_data.append(cell.value)

# D 열 앞에 열 삽입
ws.insert_cols(3, 1)
for idx, value in enumerate(ac_data, start=1):
    ws.cell(row=idx, column=3).value = value

# AC 열 삭제
ws.delete_cols(30, 1)

# 5단계: F, G 열 삭제
print("5단계: F, G 열 삭제 중...")
ws.delete_cols(6, 2)

# 6단계: J, K, L, M 열 삭제
print("6단계: J, K, L, M 열 삭제 중...")
ws.delete_cols(8, 4)

# 파일 저장
output_file = "output/reg_output.xlsx"
wb.save(output_file)
print(f"\n완료! 파일: {output_file}")
print("작업 완료")
